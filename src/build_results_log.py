#!/usr/bin/env python3
"""
Build results_master_log.xlsx — one master, tidy table for EVERY test in this project,
across all test types (forecasting, event study, regression, correlation, baseline),
and extensible to any future test. Re-run to regenerate; add rows to ROWS to log a new test.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "results_master_log.xlsx"

COLS = ["ID", "Date", "Category", "Test / Experiment", "Dataset", "Freq", "Split / Setup",
        "Model / Method", "Protocol", "Metric", "Value", "Baseline value",
        "Beats baseline?", "Verdict / Key finding", "Output files", "Notes"]

# Each row: values in COLS order (ID auto-filled). "" where n/a.
R = [
 # ---- PHASE 1: price forecasting, S&P SL 20 (daily) ----
 ["2026-07-22","Price forecast","Exp1 one-step","S&P SL20","daily","80/20","ARIMA(1,1,4)","one-step","MAPE %",0.77,0.79,"Tie","Rides persistence; ties naive","results/predictions/arima_info.json","01_arima.py"],
 ["2026-07-22","Price forecast","Exp1 one-step","S&P SL20","daily","80/20","LSTM","one-step","MAPE %",2.49,0.79,"No","Lags the price","lstm_info.json","02_lstm.py"],
 ["2026-07-22","Price forecast","Exp1 one-step","S&P SL20","daily","80/20","Hybrid","one-step","MAPE %",0.77,0.79,"Tie","Best but marginal over ARIMA","hybrid_info.json","03_hybrid.py"],
 ["2026-07-23","Price forecast","Exp2 multi-step","S&P SL20","daily","80/20 whole-window","ARIMA","multi-step (recursive)","MAPE %",33.32,33.23,"No","Flat line; misses trend","ms_arima_info.json","05"],
 ["2026-07-23","Price forecast","Exp2 multi-step","S&P SL20","daily","80/20 whole-window","LSTM","multi-step (recursive)","MAPE %",35.93,33.23,"No","Worst multi-step","ms_lstm_info.json","06"],
 ["2026-07-23","Price forecast","Exp2 multi-step","S&P SL20","daily","80/20 whole-window","Naive","multi-step (flat)","MAPE %",33.23,33.23,"—","Best multi-step (nothing beats it)","naive_baseline_info.json",""],
 ["2026-07-27","Price forecast","Exp3","S&P SL20","daily","80/20","XGBoost","one-step / multi",1.01,"",0.79,"No","1s decent; ms worst (38.8%), trees can't extrapolate","xgb_info.json","09"],
 ["2026-07-27","Price forecast","Exp3","S&P SL20","daily","80/20","GRU (3 seeds)","one-step / multi","MAPE % 1s",2.58,0.79,"No","≈ LSTM; ms 34.1%","gru_info.json","10"],
 ["2026-07-24","Price forecast","Split test","S&P SL20","daily","50/50 whole-window","ARIMA vs LSTM","multi-step","MAPE % (A/L)","30.11 / 30.06",30.50,"No","Tie; LSTM 'win' was seed noise","split5050_comparison.csv","14"],
 ["2026-07-24","Price forecast","Split test","S&P SL20","daily","40/60 whole-window","ARIMA vs LSTM","multi-step","MAPE % (A/L)","24.42 / 19.97",24.42,"No","LSTM robustly beats ARIMA ms (positioning, not skill)","split4060_comparison.csv","15"],
 ["2026-07-24","Price forecast","Paper replication","S&P SL20","daily","train2015-16 / test2017-18","ARIMA/LSTM/Hybrid","multi-step","MAPE %","6.6-6.8","~7 (paper)","Tie","Reproduces paper on calm window","(scratch check)",""],
 ["2026-07-26","Price forecast","Transformer benchmark","S&P SL20","daily","80/20/50/50/40/60","Informer/Autoformer/PatchTST/TFT","A whole / B 60-day","MAPE % (B best)","6.9-14","~6.3 (ARIMA)","No","No transformer beats ARIMA/naive (univariate)","master_transformer_vs_classical.csv","16-18"],
 ["2026-07-25","Price forecast","Forward test","S&P SL20","daily","train<=2026-04-09 / new data","ARIMA/LSTM/Hybrid","one-step / multi","MAPE % 1s",0.49,0.50,"Tie","Fresh out-of-sample; naive best multi-step; ~55pt source offset","forward2026_comparison.csv","13"],
 ["2026-07-28","Method study","Recursive vs Direct vs Multi","HNB","hourly","80/20 walk-forward","LSTM (3 methods)","fixed horizons 1/5/25/110","MAPE % @1mo (Rec/Dir)","29.3 / 13.6","5.0 (naive)","No","Direct halves recursive error at long horizon; naive still best","multistep_methods.csv","22"],
 # ---- HNB single-stock ----
 ["2026-07-26","Price forecast","HNB EOD test","HNB","daily (88 pts)","80/20","ARIMA/LSTM/Hybrid","one-step / multi","MAPE % 1s",0.31,0.31,"Tie","ARIMA=(0,1,0)=random walk=naive; LSTM worst (too little data)","hnb_eod_comparison.csv","20"],
 ["2026-07-29","Price forecast","HNB hourly test","HNB","hourly (7528)","80/20","ARIMA/LSTM/Hybrid","one-step / multi","MAPE % 1s",0.33,0.33,"Tie","Same lessons at hourly; more data didn't fix LSTM one-step","hnb_hourly_comparison.csv","21"],
 # ---- Baselines ----
 ["2026-07-29","Baseline","Prophet baseline","10 CSE stocks","daily","85/15 + MLflow","Prophet","level + direction","MAE ratio vs naive","6-87x worse","naive","No","Fails price (6-87x worse) & direction (loses to majority on 9/10)","prophet_summary_all10.csv","prophet_mlflow_all.py"],
 # ---- Hypothesis / statistical tests ----
 ["2026-07-29","Correlation","Group correlation","7 CSE stocks","daily returns","full-sample (2901 days)","Pearson corr","—","Banks-vs-Finance corr",0.251,"within 0.51/0.46","—","Banks & finance are SEPARATE clusters -> hypothesis worth pursuing","group_correlation_matrix.csv","group_correlation_test.py"],
 ["2026-07-30","Event study","Policy-rate sign-flip pilot","8 CSE stocks + ASPI","daily->event","9 hike events, CAR windows","Abnormal CAR + t/MW","event study","interaction p","n.s.","—","Null","NULL — banks & finance move SAME way on hikes; too few events","results/pilot/pilot_summary.md","pilot_policy_rate_signflip.py"],
 ["2026-07-30","Regression","Spread regression","8 CSE stocks + rates","monthly (173)","full-sample + interaction","OLS panel (cluster SE)","d_spread x is_bank","interaction p",0.0002,"—","Refuted (real but opposite)","Sig interaction BUT sign OPPOSITE hypothesis: banks -, finance +","spread_regression_summary.md","spread_regression_test.py"],
 ["2026-07-30","Regression","Spread grouping check","9 CSE stocks + rates","monthly (173)","4 specifications","OLS panel","d_spread x is_g1 (spec B)","interaction p",0.0021,"—","Survives","Effect survives dropping LOLC conglomerate; banks-negative is robust core","grouping_check_summary.md","spread_grouping_check.py"],
 ["2026-07-31","Regression","ARDL/VECM spread test","9 CSE stocks + rates","monthly (171)","ARDL bounds + VECM","ARDL/VECM","long-run vs short-run","ARDL bounds F (banks)",0.66,"—","No","No long-run link (spread is I(0)); effect is short-run only; banks -, finance + but weak long-run","ardl_vecm_summary.md","spread_ardl_vecm.py"],
 ["2026-07-31","Direction","Direction Step 1 baseline","HNB","daily","80/20 chrono","XGBoost","next-day Buy/Sell/Hold","Accuracy %",36.2,43.1,"No","Price+technical only: no edge, loses to persistence (43%). Honest baseline to beat.","step1_summary.md","direction_step1_baseline.py"],
 # ---- PHASE A: multi-horizon direction + return (Tier-1 technical only) ----
 ["2026-08-01","Direction","Phase A multi-horizon direction","HNB","daily","80/20 chrono + h-bar purge","Logistic + XGBoost","direct, 9 horizons (1d-252d)","Horizons beating baseline","0 of 9","majority / persistence","No","NO EDGE AT ANY HORIZON. Best model always below best baseline; gap widens with horizon (-3.5pp @1d -> -47.1pp @1yr) because long-horizon drift makes the dumb baseline huge (maj 78% @6mo, pers 77% @1yr).","results/direction/multi_horizon/mh_summary.md","direction_multi_horizon.py; Tier-1 only; dead-zone +/-0.5%*sqrt(h)"],
 ["2026-08-01","Direction","Phase A best single horizon","HNB","daily","80/20 chrono + h-bar purge","Logistic","direct, h=22d (1 month)","Accuracy %",40.1,42.1,"No","Best model result of the sweep (1-month Logistic 40.1%) still loses to persistence 42.1%. Logistic > XGBoost at every horizon -> no nonlinear structure to find.","mh_direction_table.csv","edge -2.0pp"],
 ["2026-08-01","Return forecast","Phase A multi-horizon return %","HNB","daily","80/20 chrono + h-bar purge","Ridge + XGBoost","direct, 9 horizons (1d-252d)","RMSE ratio vs train-mean",0.99,1.0,"Tie","Ridge beats naive-0 at 8/9 horizons BUT is only ~1% better than train-mean drift -> it learned drift, NOT signal. XGBoost worse than naive everywhere (ratio 1.01-1.14).","mh_return_table.csv","train-mean is the honest null, not naive-0"],
 ["2026-08-01","Return forecast","Phase A return sign accuracy","HNB","daily","80/20 chrono + h-bar purge","Ridge","direct, sign of h-day return","Horizons beating sign null","2 of 9","always-guess-winning-side","No","TRAP CHECK: 55-57% sign accuracy @1-3mo looks like the 52-58% target but 'always guess up' matches/beats it (test window drifted up 59-95%). Both 'wins' are +0.2pp = noise.","mh_edge_vs_horizon.png","up-share null added to stop false positive"],
 # ---- PHASE B: ablation, + Tier-2 technical (RSI, MACD, volume) ----
 ["2026-08-01","Direction","Phase B ablation (+Tier-2)","HNB","daily","identical rows/split/seeds, A vs B","Logistic + XGBoost","direct, 9 horizons, ablation","Mean gain from Tier-2 (pp)",-0.9,0,"No","RSI+MACD+volume add NOTHING. Mean -0.9pp, 0/9 beat baseline. Technical indicators exhausted.","results/direction/phase_ablation/ablation_summary.md","direction_phase_ablation.py; 9 -> 16 features; value restated on the 4-phase run (macro availability trims early rows)"],
 ["2026-08-01","Direction","Phase B feature attention check","HNB","daily","identical rows/split/seeds","XGBoost importance","direct, 9 horizons","Tier-2 share of importance %",41,"—","—","XGBoost spends 41% of importance on Tier-2 (macd_signal is top feature at 6/9 horizons) yet accuracy does not improve -> the model USES them but they carry no signal.","ablation_tier2_importance.csv","strong evidence indicators are redundant with price"],
 ["2026-08-01","Return forecast","Phase B return % (+Tier-2)","HNB","daily","identical rows/split/seeds","Ridge + XGBoost","direct, 9 horizons","RMSE ratio vs train-mean (B)",0.991,0.991,"Tie","Tier-2 made return RMSE slightly WORSE at 9/9 horizons (ratio gain +0.000 to +0.046). Sign edge stays negative everywhere.","ablation_return_table.csv","A ratio 0.991 -> B ratio 0.991-1.034"],
 # ---- PHASE C: ablation, + macro interest rates (levels + changes), with C2 diagnostic ----
 ["2026-08-01","Direction","Phase C ablation (+macro rates)","HNB","daily","identical rows/split/seeds, 35d publication lag","Logistic + XGBoost","direct, 9 horizons, ablation","Mean gain from macro (pp)",-2.8,0,"No","Macro rates make it WORSE and unstable: gain swings -22.3 to +10.3 pp, 0/9 beat baseline. Not weak signal - overfitting.","results/direction/phase_ablation/ablation_summary.md","direction_phase_ablation.py; 16 -> 29 features; inflation/FX/M2 still uncollected"],
 ["2026-08-01","Direction","Phase C2 diagnostic (macro changes only)","HNB","daily","identical rows/split/seeds","Logistic + XGBoost","direct, 9 horizons, ablation","Mean gain from macro d (pp)",3.4,-2.8,"—","DROPPING RATE LEVELS FIXES THE DAMAGE (-2.8 -> +3.4 pp; 1yr 4.5% -> 49.7%). Non-stationary levels were being memorised. Still 0/9 beat baseline -> form was wrong AND no signal.","ablation_gain_table.csv","key methodology lesson: feed macro as changes, never levels"],
 ["2026-08-01","Direction","Phase C feature attention check","HNB","daily","identical rows/split/seeds","XGBoost importance","direct, 9 horizons","Macro share of importance %",64,"—","—","Macro importance rises 46% (1d) -> 80% (1yr) while accuracy FALLS. Model fits rates as a trend proxy, not signal. tb_12m/spread/policy_rate are top features at 8/9 horizons.","ablation_feature_importance.csv","fingerprint of trending-variable overfit"],
 ["2026-08-01","Return forecast","Phase C return % (+macro)","HNB","daily","identical rows/split/seeds","Ridge + XGBoost","direct, 9 horizons","RMSE ratio vs train-mean (C)",1.37,0.995,"No","Macro blows out return RMSE to 1.29-1.37x the train-mean null (A/B sat at ~0.99). Clear damage, confirms the overfit reading.","ablation_return_table.csv","worst at 2-3 month horizons"],
 # ---- PHASE D: ablation, + sector (ASPI, peer banks/finance, relative strength, beta) ----
 ["2026-08-01","Direction","Phase D ablation (+sector)","HNB","daily","identical rows/split/seeds","Logistic + XGBoost","direct, 9 horizons, ablation","Mean gain from sector (pp)",2.6,0,"No","FIRST consistently POSITIVE step: +2.6pp mean, positive at 6/9 horizons. Built on C2 (macro changes only). But still 0/9 beat baseline.","results/direction/phase_ablation/ablation_summary.md","direction_phase_ablation.py; C2 16+7 -> D 39 features; ASPI + COMB/SAMP + LOFC/LOLC/LFIN/CFIN"],
 ["2026-08-01","Direction","Phase D best single horizon","HNB","daily","identical rows/split/seeds","Logistic + XGBoost","direct, h=5d (1 week)","Edge over baseline (pp)",-0.2,0,"No","Closest call of the whole study: 1-week Phase D 37.3% vs baseline 37.4%. Near-parity at short horizons (1d -1.7, 1wk -0.2, 2wk -1.6) but never crosses.","ablation_gain_table.csv","not a win; single split, one stock"],
 ["2026-08-01","Return forecast","Phase D return sign edge (+sector)","HNB","daily","identical rows/split/seeds","Ridge + XGBoost","direct, sign of h-day return","Sign edge 1wk/2wk/3wk (pp)","+1.7 / +1.7 / +1.1","always-guess-winning-side","Tie","First POSITIVE sign edges in the study, and at 3 adjacent horizons (1-3 weeks). Tiny (~1-2pp) and one split only -> flag as worth re-testing, NOT a finding.","ablation_return_table.csv","check on other tickers before claiming anything"],
 ["2026-08-01","Direction","Phase D feature attention check","HNB","daily","identical rows/split/seeds","XGBoost importance","direct, 9 horizons","Sector share of importance %",34,"—","—","Sector takes 34% avg (42% at 1 day, falling to 27% at 1yr) - mirror image of macro, which rises with horizon. Sector matters short, macro long, neither enough.","ablation_feature_importance.csv","vol_20 is top feature at 1d/1wk; d_tb3m dominates long"],
 # ---- SECTOR SWEEP: replication of Phase D across 11 stocks (banks/finance/control) ----
 ["2026-08-01","Direction","Sector sweep - replication test","11 CSE stocks (3 banks, 4 finance, 4 control)","daily","80/20 chrono + h-bar purge, per stock","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Cells beating baseline","27 of 99","14 of 99 (Phase A)","No","Sector features nearly DOUBLE the win rate vs Tier-1 (14->27 of 99) but wins sit at exactly coin-flip at short horizons and collapse at long ones. No real edge.","results/direction/sector_sweep/sweep_summary.md","direction_sector_sweep.py; peer composites exclude the target stock"],
 ["2026-08-01","Direction","Sector gain replication (HNB claim 3)","11 CSE stocks","daily","identical protocol per stock","Phase D - Phase A","direct, 9 horizons x 11 stocks","Mean sector gain (pp)",3.8,0,"—","REPLICATES: +3.8pp mean, positive in 66/99 cells (67%), 9 of 11 stocks positive. HNB was not a fluke - sector context genuinely helps the model. Best: LFIN +13.2, SAMP +9.4, LOFC +8.2. Worst: CTC -11.0.","sweep_gain_by_ticker.csv","helping is not the same as winning"],
 ["2026-08-01","Direction","Sector sweep significance screen","11 CSE stocks","daily","sign test + Wilcoxon across stocks, per horizon","binomial sign test (H0 p=0.5)","direct, 9 horizons","Best sign-test p",0.5,0.05,"No","NOTHING SIGNIFICANT AT ANY HORIZON. Best cells (1 day, 1 week) are 6/11 stocks positive = p=0.500, exactly a coin flip. Long horizons far worse (1yr 0/11).","sweep_significance.csv","stocks share a market factor so p-values are already optimistic"],
 ["2026-08-01","Direction","HNB sign-edge claim retest (claim 2)","11 CSE stocks","daily","identical protocol per stock","Ridge + XGBoost","direct, sign of h-day return, 1-3 week horizons","Cells with positive sign edge","6 of 33","~17 of 33 (coin flip)","No","REFUTED. HNB's positive 1-3 week sign edge did NOT replicate - 18% vs the ~50% chance would give. It was luck. Flagged as provisional when found, now closed.","sweep_all_results.csv","value of naming a claim before retesting it"],
 # ---- PHASE E: daily macro (was monthly frequency the problem?) ----
 ["2026-08-01","Direction","Phase E - daily USD/LKR","11 CSE stocks","daily","80/20 chrono + h-bar purge, 2015-2026 window","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Mean gain from daily FX (pp)",0.0,0,"No","FREQUENCY WAS NOT THE PROBLEM. Daily FX gain is exactly 0.0pp (positive in 57/99 cells = coin flip). Macro fails daily exactly as it failed monthly.","results/direction/daily_macro/daily_macro_summary.md","direction_daily_macro.py; new cleaned_data/usd_lkr_daily.csv"],
 ["2026-08-01","Direction","Phase E2 - daily global + CPI","11 CSE stocks","daily","identical rows/split/seeds","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Mean gain from global+CPI (pp)",-1.9,0,"No","Oil, US 10Y, DXY and inflation make it WORSE (-1.9pp), worst at long horizons (-6.0pp at 1yr). Same trending-variable overfit as Phase C.","daily_macro_gain.csv","CPI spliced across 4 base years; money supply excluded (ends 2024-08, inside test window)"],
 ["2026-08-01","Direction","Phase E significance screen","11 CSE stocks","daily","sign test across stocks, 4 phases x 9 horizons","binomial sign test (H0 p=0.5)","direct","Combinations with p<0.05","0 of 36",0.05,"No","NOTHING SIGNIFICANT ANYWHERE. Best cell is 1 day (6/11 stocks, p=0.500) - exact coin flip, same as the sector sweep. Cells beating baseline: A 17, D 27, E 24, E2 25 of 99.","daily_macro_significance.csv","macro information source now closed at both frequencies"],
 ["2026-08-01","Direction","Phase E feature attention check","11 CSE stocks","daily","identical rows/split/seeds","XGBoost importance","direct, 9 horizons x 11 stocks","New-data share of importance %",27,"—","—","Model spends 27% of importance on the new data (FX 7%, global 11%, CPI 9%) and gains nothing - third time this fingerprint has appeared (Tier-2 41%, macro 64%, now 27%).","daily_macro_all_results.csv","using a feature is not the same as profiting from it"],
 # ---- PHASE F: money supply + full classic macro set, on a MOVED (crisis-era) window ----
 ["2026-08-01","Direction","Phase F - money supply (M1/M2/M2b/M4)","11 CSE stocks","daily","80/20 chrono, sample cut at 2024-08; test 2022-03 to 2024-08","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Mean gain from money supply (pp)",-0.7,0,"No","Money supply adds NOTHING even on a crisis-era test window (positive in 50/99 cells = exact coin flip). Window moved back so M1/M2 could be tested fairly rather than excluded.","results/direction/money_supply/money_summary.md","direction_money_supply.py; 65-day publication lag per CBSL's own note; reserve money dropped (ends 2024-02)"],
 ["2026-08-01","Direction","Phase F - full classic macro set","11 CSE stocks","daily","identical rows/split/seeds","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Mean gain money->full macro (pp)",1.1,0,"No","Full classic set (money + rates + FX + inflation) = 4 of the 5 Naik-Padhi variables. Still 42/99 cells, below the money-only phase (51/99). Only industrial production remains uncollected.","money_gain.csv","closes the macro chapter apart from IIP"],
 ["2026-08-01","Direction","Crisis-window effect (window sensitivity)","11 CSE stocks","daily","2022-03 to 2024-08 vs 2024-2026","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Cells beating baseline","43-51 of 99","17-27 of 99 (calm window)","—","IMPORTANT METHOD FINDING: moving the test window into the 2022 crisis roughly DOUBLES the win rate for ALL phases including price-only. Results are window-dependent, not model-dependent.","money_by_horizon.csv","argues for walk-forward / multi-window evaluation in the write-up"],
 ["2026-08-01","Direction","1-day significance - macro ruled out","11 CSE stocks","daily","sign test across stocks, 4 phases x 9 horizons","binomial sign test (H0 p=0.5)","direct","Cells with p<0.05","2 of 36","1.8 expected by chance","No","Both hits are at 1 day, and the STRONGEST is Phase A (price only, 9/11, +3.9pp, p=0.033). Adding money supply SHRINKS it to +3.2pp. So it is the crisis window, not macro - and 2/36 is chance anyway.","money_significance.csv","checking the null model is what prevented a false claim"],
 # ---- PHASE G: industrial production (IIP) - closes the classic macro set ----
 ["2026-08-02","Direction","Phase G - industrial production (IIP)","11 CSE stocks","daily","80/20 chrono, sample 2016-2026 (IIP coverage)","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Mean gain from IIP (pp)",-6.4,0,"No","IIP makes it WORSE (-6.4pp, positive in only 25/99 cells, 0/36 significant). It was the STRONGEST variable in the Naik-Padhi reference paper, and it still does not predict direction.","results/direction/industrial_production/iip_summary.md","direction_industrial_production.py; DCS index entered by hand, 2016-01 to 2026-05; 50-day publication lag"],
 ["2026-08-02","Direction","Phase G - full classic macro set complete","11 CSE stocks","daily","identical rows/split/seeds","Logistic + XGBoost","direct, 9 horizons x 11 stocks","Cells beating baseline","25 of 99","27 of 99 (no macro)","No","ALL FIVE Naik-Padhi macro variables now tested (IIP, money supply, T-bill rate, exchange rate, inflation). NONE beats plain price history. Macro chapter complete.","iip_gain.csv","adding all macro is WORSE than sector-only (27/99)"],
 ["2026-08-02","Data quality","IIP seasonality guard","Sri Lanka DCS IIP","monthly (125)","2016-01 to 2026-05","seasonal inspection","n/a","April vs March gap (pts)","-10 to -34","—","—","IIP is NOT seasonally adjusted - April falls 10-34 pts EVERY year (Sinhala/Tamil New Year), March always peaks. Month-on-month would be a calendar signal, so YoY-only features were used.","cleaned_data/industrial_production_monthly.csv","clean_macro_sources.py; mom column deliberately named _DO_NOT_USE"],
 # ---- POOLED PANEL + CO-MOVEMENT + READABLE RULES ----
 ["2026-08-02","Correlation","Co-movement lift table (no model)","7 bank/finance stocks (22,677 stock-days)","daily","full panel, 15 indicators x 9 horizons","conditional probability + 2-proportion z","direct","Median |lift| short horizons (pp)",2.5,0,"—","'If indicator up, does stock go up?' answered by pure counting. Market/peer indicators give a real +6 to +10pp lift at 1d-1mo on huge samples. Macro lifts are near zero at short horizons.","results/direction/pooled_rules/comovement_table.csv","direction_pooled_and_rules.py; lift = P(up|ind up) - P(up|ind down)"],
 ["2026-08-02","Correlation","Co-movement - long-horizon lifts are artifacts","7 bank/finance stocks","daily","panel, |lift|>=10pp cells","independent-window audit","direct","Independent windows behind biggest lifts","7-24","hundreds needed","No","All 19 large lifts (T-bill -19.3, M2 +17.0, USD/LKR -15.3) sit at 3mo-1yr where only 7-24 INDEPENDENT windows exist. Overlapping windows inflate the z-test. Discard them.","comovement_table.csv","median |lift| long 4.2pp vs short 2.5pp is a sample-size illusion"],
 ["2026-08-02","Direction","Pooled panel model (sector-wide)","7 bank/finance stocks pooled","daily","split by DATE not row, 17,770 train rows","Logistic + XGBoost","direct, 9 horizons, one model for whole sector","Horizons beating pooled baseline","2 of 9","majority / persistence","No","7x more training data (17,770 vs ~2,700 per stock). Beats the POOLED baseline at 1 day (+3.3pp) and 1 week (+2.7pp) only.","results/direction/pooled_rules/pooled_model_results.csv","direction_pooled_and_rules.py"],
 ["2026-08-02","Direction","Pooled model fairness check","7 bank/finance stocks","daily","pooled model scored vs each stock's OWN baseline","binomial sign test (H0 p=0.5)","direct, 9 horizons","1-day: stocks positive","6 of 7","p=0.062","No","STRONGEST RESULT OF THE PROJECT but still not significant. The 1-week +2.7pp was a POOLING ARTIFACT (-0.5pp vs own baselines). Only 1 day survives: median +2.5pp, p=0.062, 1 of 9 horizons, single window.","pooled_vs_own_baseline.csv","pooling stocks with different class balance weakens the majority baseline and fakes an edge"],
 ["2026-08-02","Direction","Readable rules extracted","7 bank/finance stocks pooled","daily","depth-3 tree, scored on unseen test set","Decision tree (depth 3)","direct, h=1/5/22","Best 1-day rule accuracy %",49.3,38.0,"Yes","17 plain-English rules with out-of-sample hit rates. Best 1-day rule: calm market + flat ASPI + no big drop -> Hold, 49.3% (+11.3pp). Shows WHAT the model believes, not just that it fails.","readable_rules.csv","1-month rule at 68.5% fires only 317 times - treat as noise"],
 # ---- WALK-FORWARD RE-TEST OF THE 1-DAY CLAIM ----
 ["2026-08-02","Direction","1-day claim re-test (walk-forward)","7 bank/finance stocks","daily","21 non-overlapping 6-month windows, expanding train","Logistic + XGBoost (reported separately)","direct, walk-forward, h=1","Folds with positive median edge","6-7 of 21","10.5 expected by chance","No","CLAIM REFUTED. Median of fold medians -1.8pp (Logistic) / -1.9pp (XGBoost), sign-test p=0.96 and 0.99. The single-split +2.5pp did not survive 21 independent windows.","results/direction/retest_1day/retest_summary.md","direction_1day_retest.py; verdict rule fixed IN ADVANCE; 1-week carried as control, also negative"],
 ["2026-08-02","Direction","1-day re-test - window regime effect","7 bank/finance stocks","daily","21 walk-forward folds","Logistic + XGBoost","direct, h=1","Positive folds 2016-2020 vs 2021-2024","1 of 10 vs 6 of 8","—","—","Confirms the crisis-window finding a second time: folds in 2021-2024 are mostly POSITIVE, 2016-2020 mostly negative. Apparent skill is a property of the period, not the model.","retest_by_fold.csv","strong argument for walk-forward reporting in the dissertation"],
 ["2026-08-02","Direction","1-day re-test - per stock consistency","7 bank/finance stocks","daily","21 walk-forward folds each","XGBoost","direct, h=1","Stocks positive in >50% of folds","2 of 7","—","No","No consistency: LOFC and CFIN positive in 62% of folds, HNB in only 19%. A real edge would show across the sector, not in two names.","retest_by_stock.csv","median edge per stock ranges +4.9 (LOFC) to -4.9 (SAMP)"],
]

# ---- write workbook ----
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Results log"
hdr_fill = PatternFill("solid", fgColor="1F3864"); hdr_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="D0D0D0"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(COLS)
for c in range(1, len(COLS) + 1):
    cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
verdict_fill = {"Yes": "C6EFCE", "No": "FFC7CE", "Tie": "FFEB9C", "—": "F2F2F2"}
for i, row in enumerate(R, start=1):
    ws.append([i] + row)
    r = i + 1
    beats = str(ws.cell(r, 13).value)
    if beats in verdict_fill:
        ws.cell(r, 13).fill = PatternFill("solid", fgColor=verdict_fill[beats])
    ws.cell(r, 13).alignment = Alignment(horizontal="center")
    ws.cell(r, 14).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, len(COLS) + 1):
        ws.cell(r, c).border = border
widths = [4,11,14,22,16,10,20,22,20,16,10,12,13,42,26,14]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(COLS))}{len(R)+1}"

# ---- README sheet ----
rm = wb.create_sheet("README")
readme = [
 ["RESULTS MASTER LOG — how to use", ""],
 ["", ""],
 ["Purpose", "One tidy row per reported test result, across ALL test types (forecasting, regression, event study, correlation, baseline). Filter/sort by any column."],
 ["", ""],
 ["Column", "Meaning"],
 ["ID", "Sequential row number"],
 ["Date", "When the test was run (approx.)"],
 ["Category", "Price forecast / Baseline / Correlation / Event study / Regression / Method study / (new types allowed)"],
 ["Test / Experiment", "Short test name"],
 ["Dataset", "What data (S&P SL20, HNB, 10 CSE stocks, ...)"],
 ["Freq", "daily / hourly / monthly / event"],
 ["Split / Setup", "80/20, 50/50, 40/60, walk-forward, full-sample, event-study, n specifications, ..."],
 ["Model / Method", "ARIMA, LSTM, Hybrid, Prophet, XGBoost, GRU, TFT, OLS, event-study, ..."],
 ["Protocol", "one-step / multi-step / recursive / direct / n/a"],
 ["Metric", "Name of the reported number (MAPE %, accuracy, interaction p, CAR, correlation, ...)"],
 ["Value", "The metric value"],
 ["Baseline value", "The naive / comparison value (leave '—' if not applicable)"],
 ["Beats baseline?", "Yes / No / Tie / —  (drives the colour)"],
 ["Verdict / Key finding", "One-line human takeaway"],
 ["Output files", "Where the detailed results live"],
 ["Notes", "Anything else (caveats, script name)"],
 ["", ""],
 ["RULE", "Update this log after EVERY test run (this chat or Claude Code): add a new row to ROWS in src/build_results_log.py and re-run it, OR append a row directly in Excel. Never leave a completed test unlogged."],
 ["To regenerate", "python src/build_results_log.py"],
]
for row in readme:
    rm.append(row)
rm.cell(1, 1).font = Font(bold=True, size=14)
rm.cell(5, 1).font = Font(bold=True); rm.cell(5, 2).font = Font(bold=True)
for r in range(6, 21):
    rm.cell(r, 1).font = Font(bold=True)
rm.cell(23, 1).font = Font(bold=True, color="C00000")
rm.column_dimensions["A"].width = 22; rm.column_dimensions["B"].width = 95
for r in range(1, len(readme) + 1):
    rm.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print(f"Saved {OUT}  ({len(R)} results, {len(COLS)} columns)")
